import logging
import pprint
import re
from collections import deque

from getrealty import printArray, sqlConnect, settings, mydirs
import openpyxl
import xlsxwriter

pp = pprint.PrettyPrinter(width=1)
logger = logging.getLogger(__name__)
config = settings.config


def writeExcel(rnumbers):
    """Library for writing properties to an Excel Spreadsheet"""

    num_rnumbers = len(rnumbers)
    rnumber_arrays = []
    # build arrays of rnumbers arrays with max in each of from defined
    # DEFAULT value
    if config['defaults']['SEPARATE_WKTS'] is True:
        pass
        # it = natatime config.DEFAULT_MAX_PER_SPREADSHEET, rnumbers_only
        # while (my @vals = $it->()) {
        # 	push(@rnumber_arrays, \@vals);
    else:
        rnumber_arrays.append(rnumbers)

    # my $num = $#rnumbers_only + 1;

    num_rnum_arrays = len(rnumber_arrays)
    logger.info("Writing %s Rnumbers", num_rnum_arrays)

    i = 0
    excel_dir = mydirs.MyDirs().exceldir()
    def_output = config['defaults']['OUTPUT']
    # writefile = config.default_outputxls
    writefile = excel_dir + '/' + def_output + ".xlsx"
    print("writefile: ", writefile)
    workbook = xlsxwriter.Workbook(writefile)

    # print(rnumber_arrays)
    for array in rnumber_arrays:

        # there is only one file, don't split unless told to
        if num_rnum_arrays is not 0:
            pass
            # writefile =~ s/(.*)(.xls.?)/$1.$i$2/;

        if config['defaults']['NO_QUERY_OVERWRITE_XLS'] is True:
            pass
            # questionUserIfWantToOverWriteXLS(writefile)

        writeWorkBook(workbook, num_rnumbers, array)

        i += 1

    workbook.close()

    wb = openpyxl.load_workbook(filename=writefile)
    worksheet = wb.active

    # this works, but adds a lot of space, still haven't found good replacement
    # for perl autofit routine
    # https://stackoverflow.com/questions/39529662/python-automatically-adjust-width-of-an-excel-files-columns
    # my_autofit_columns(worksheet)
    wb.save(writefile)


def writeWorkBook(workbook, num_rnumbers, rnumbers_only):

    myPA = printArray.MyPrintArray()
    printhash = myPA.CreatePrintHash()

    worksheet = workbook.add_worksheet("Summary")
    worksheet_db = workbook.add_worksheet("Summary_db")
    worksheet_rnum = \
        workbook.add_worksheet(config['defaults_static']['RNUM_NOTES_SHEET'])

    # worksheet.add_write_handler(qr[\w], \&store_string_widths);
    # worksheet_db.add_write_handler(qr[\w], \&store_string_widths);
    # worksheet_rnum.add_write_handler(qr[\w], \&store_string_widths);

    row = 0
    col = 1

    # write the merged headers
    (row, col) = writeMergedHeaders(workbook, worksheet,
                                    worksheet_db, row, col, printhash)

    col = 0

    # Write the headings
    # get non-merged headers
    (row, col) = writeNonMergedHeaders(myPA, workbook, worksheet, worksheet_db,
                                       row, col, printhash)
    # exit()

    row_rnum_sheet = 0
    col_rnum_sheet = 0

    # cycle through rnumbers
    # sql query cannot handle more than 1000 at a time
    if len(rnumbers_only) > 900:
        pass
    # 	my $it = natatime 900, @rnumbers_only;
    # 	while (my @vals = $it->()) {
    #        ($row,$col) = writeRnumberRows($workbook,$worksheet,$worksheet_db,
    #                       $row,$col,$printhash,@vals);
    # ($row_rnum_sheet,$col_rnum_sheet) = writeRnumberSheet($workbook,
    #                                              $worksheet_rnum,
    #                                              $row_rnum_sheet,
    #                                              $col_rnum_sheet,
    #                                              $num_rnumbers,
    #                                              $printhash,@vals);
    else:
        (row, col) = writeRnumberRows(myPA, workbook, worksheet, worksheet_db,
                                      row, col, printhash, rnumbers_only)
        # ($row_rnum_sheet,$col_rnum_sheet) = writeRnumberSheet($workbook,
        #                                                       $worksheet_rnum,
        #                                                       $row_rnum_sheet,
        #                                                       $col_rnum_sheet,
        #                                                       $num_rnumbers,
        #                                                       $printhash,
        #                                                       @rnumbers_only);

    worksheet.freeze_panes(2, 3)  # Freeze row, col
    worksheet_db.freeze_panes(2, 3)  # Freeze row, col

    headingsArray = myPA.getHeadingsFromPrintArray(1, 1)
    num_headers = len(headingsArray)

    # VERY IMPORTANT -- If you dont auto filter on all the columns, the sort
    #       will only sort the columns in the autosort range.
    #       This will leave columns to the left unchanged, while rearranging
    #       all the columns to the right which are in the sort range
    worksheet.autofilter(1, 0, num_rnumbers + 1, num_headers + 1)
    worksheet_db.autofilter(1, 0, num_rnumbers + 1, num_headers + 1)
    # $worksheet->filter_column( 'A', 'x =~ R*' );
    # autofit_columns($worksheet);
    # autofit_columns($worksheet_db);
    # autofit_columns($worksheet_rnum);
    #
    # #jafhide$worksheet_db->hide();


def my_autofit_columns(worksheet):

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
    adjusted_width = (max_length + 2) * 1.2
    worksheet.column_dimensions[column].width = adjusted_width


def writeRnumberRows(myPA, workbook, worksheet, worksheet_db, row, col,
                     printhash, rnumbers_only):

    # get the desired Rnumbers from the database
    # to build the select statement
    rnumbers_only_quoted = []
    for rnumber in rnumbers_only:
        rnumbers_only_quoted.append('"' + rnumber + '"')

    table_name = config['defaults']['TABLE_NAME']
    # print("table_name", table_name)

    # SELECT desired Rnumber rows and get all values from db
    sql_where = "SELECT * FROM {} WHERE ".format(table_name)
    sql_where = sql_where + "r_num=" + " OR r_num=".join(rnumbers_only_quoted)

    logger.info("sql_where = %s", sql_where)

    work_dir = config['defaults']['WORK_DIR']
    sql_file = work_dir + '/RealEstate.db'

    table_values = sqlConnect.sqlQueryRequest(sql_where)

    # get all the headings from the db
    sql_query = 'PRAGMA table_info( {} )'.format(table_name)
    sql_results = sqlConnect.sqlQueryRequest(sql_query)

    db_headings = [tup[1] for tup in sql_results]

    # contains info for rnumbers (created by WHERE)
    # store in a hash.  Going to get values and order of
    # columns from the printArray. Before got them from db, but couldn't
    # reorder columns easily
    hash_values = {'rnum_values': {}, 'rnumbers': []}
    for row_table in table_values:
        col = 0
        tmp_hash = dict()
        for db_heading in db_headings:
            col_letter = ntol(col + 1)
            row_p1 = row + 1
            value = row_table[col]
            tmp_hash[db_heading] = value
            if re.match("r_num", db_heading):
                rnum = value
                # hash_values = ['rnum_values'].update({rnum:{}})
                hash_values['rnum_values'].update({rnum: {}})
            col += 1
            # print("tmp_hash\n",tmp_hash)
        hash_values['rnum_values'][rnum].update(tmp_hash)
        hash_values['rnumbers'].append(rnum)

    # print(hash_values['rnum_values'])
    # get the columns to print.  column order is from printArray
    headingsArray = ['r_num'] + \
        myPA.getHeadingsFromPrintArray(1, 1)

    # pp.pprint(hash_values['rnumbers'])
    # pp.pprint(hash_values['rnum_values'])

    # print the rnumbers
    for rnumber in hash_values['rnumbers']:
        # print("rnumber", rnumber)

        col = 0
        cnt = 0
        for heading in headingsArray:
            cnt += 1
            # if cnt is 35:
            #   break

            col_letter = ntol(col + 1)
            row_p1 = row + 1

            value = hash_values['rnum_values'][rnumber][heading]
            # define rpg value as 1 or 0
            if heading is "rpg":
                NOTES_value = \
                    hash_values['rnum_values'][rnumber]['NOTES_rnum_sheet']
                # print("NOTES_value",NOTES_value)
                if NOTES_value is "":
                    value = 0
                else:
                    value = 1

            # BUG in excel..   If a cell is conditionally formatted, you cannot
            # add alignment!!
            my_format = workbook.add_format()
            if heading in printhash['format']:
                format_val = printhash['format'][heading]
                format_new = setFormatting(workbook, value,
                                           my_format, format_val)
                # format_new = setFormatting(workbook,
                #                            value,
                #                            my_format,
                #                            format_val)
            else:
                format_new = ""

            # conditional formatting
            if heading in printhash['cond_format']:
                cond_format = printhash['cond_format'][heading]
                setConditionalFormatting(workbook, worksheet, cond_format,
                                         col_letter, row_p1)
                # setConditionalFormatting(workbook,worksheet_db,cond_format,col_letter,row_p1)

            label = ""
            if heading in printhash['type']:
                # print("1. value = {},  heading = {}".format(value, heading))
                my_type = printhash['type'][heading]

                if my_type is "file":
                    value = "external:" + value
                    label = "file"
                elif my_type is "goog":
                    if value is not "":
                        value = getGoogleLink(value)
                        label = "map"
                    else:
                        value = "n/a"
                        worksheet.write(row, col, value,
                                        getWkbFormats(workbook,
                                                      "format_orange"))
                        worksheet_db.write(row, col, value,
                                           getWkbFormats(workbook,
                                                         "format_orange"))
                        col += 1
                        continue
                elif my_type is "gis":
                    if value:
                        value = getGoogleLink(value)
                        label = "GIS"
                elif re.match(r"link:(\S+)", my_type):
                    m = re.match(r"link:(\S+)", my_type)
                    label = m.group(1)
                new_val = str(ntol(col + 1)) + str(row + 1)
                worksheet.write_url(new_val, str(value), string=label)
                worksheet_db.write_url(new_val, value, string=label)
            else:
                # print("2. {}, {}, value = {}, heading = {}".format(col,
                #                                                    row,
                #                                                   value,
                #                                                    heading))
                print_format = getWkbFormats(workbook, "format_mobilehome")
                # print_format = (col is 0) and (value =~ /^M/) ? getWkbFormats(workbook,"format_mobilehome") :
                #                     (col is 0)                 ? getWkbFormats(workbook,"format_non_merged") :
                #                     format_new                 ? format_new :
                #                     getWkbFormats(workbook,"format_default")

                if (col is 0) and re.match('^M', value):
                    print_format = getWkbFormats(workbook, "format_mobilehome")
                elif (col is 0):
                    print_format = getWkbFormats(workbook, "format_non_merged")
                elif format_new:
                    print_format = format_new
                else:
                    print_format = getWkbFormats(workbook, "format_default")

                worksheet.write(row, col, value, print_format)
                worksheet_db.write(row, col, value, print_format)

            col += 1
        row += 1
    return(row, col)


def getGoogleLink(addr):
    """Return google maps link for given address"""

    link = "http://maps.google.com/?q=" + addr + "&sensor=true"

    return(link)


def setConditionalFormatting(workbook, worksheet, cond_format,
                             col_letter, row_p1):

    # Issues with using conditional formatting.
    # 1. Large spreadsheets will run slower
    # 2. cannot right justify format after setting conditional

    if cond_format is "format1":
        # light red for 0
        # jaf Cannot get Y and N to work

        writeCondText(worksheet, col_letter, row_p1, "begins with", "1",
                      getWkbFormats(workbook, "format_grn"))
        writeCondCell(worksheet, col_letter, row_p1, "=",
                      "0", getWkbFormats(workbook, "format_red"))
        writeCondText(worksheet, col_letter, row_p1,
                      "begins with", "1 np", getWkbFormats(workbook,
                                                           "format_grn"))
        writeCondText(worksheet, col_letter, row_p1,
                      "begins with", "0 no", getWkbFormats(workbook,
                                                           "format_red"))
        writeCondText(worksheet, col_letter, row_p1,
                      "begins with","n/a", getWkbFormats(workbook,
                                                         "format_orange"))
    elif cond_format is "format2":
        # light red for 0
        # jaf Cannot get Y and N to work
        # $worksheet->conditional_formatting( "$col$row",
        cell = col_letter + row_p1
        worksheet.conditional_formatting(cell,
                                         {
                                             'type': 'cell',
                                             'criteria': "<",
                                             'value': "5",
                                             'format': getWkbFormats(
                                                 workbook,
                                                 "format_grn")
                                         })
        worksheet.conditional_formatting(cell,
                                         {
                                            'type': 'cell',
                                            'criteria': "between",
                                            'minimum': "5",
                                            'maximum': "20",
                                            'format': getWkbFormats(
                                                workbook, "format_orange"),
                                         })
        worksheet.conditional_formatting(cell,
                                         {
                                             'type': 'cell',
                                             'criteria': ">",
                                             'value': "20",
                                             'format': getWkbFormats(
                                                 workbook, "format_red"),
                                         })
    else:
        pass
        # $logger->error("conditional format \"$cond_format\" not defined");


def setFormatting(workbook, value, my_format, format_val):

    # print ("format_val", format_val, "value", value)
    if format_val is "formatR":
        my_format.set_align('right')
    elif format_val is "format1or0":
        if re.match('^1', str(value)):
            my_format = getWkbFormats(workbook, "format_grn")
        elif re.match('^0', str(value)):
            my_format = getWkbFormats(workbook, "format_red")
        elif re.match(r'n\/a', str(value)):
            my_format = getWkbFormats(workbook, "format_orange")
        else:
            my_format = ""
        if my_format is not "":
            my_format.set_align('center')
    elif format_val is "formatpct":
        low_val = 5
        mid_val = 20
        # my_format = getWkbFormats(workbook, "format_grn")

        value = float(value)
        if value < low_val:
            my_format = getWkbFormats(workbook, "format_grn")
        elif value >= low_val and value <= mid_val:
            my_format = getWkbFormats(workbook, "format_orange")
        elif value > mid_val:
            my_format = getWkbFormats(workbook, "format_red")
        else:
            my_format = ""
        my_format.set_align('right')
    elif format_val is "lt_grey":
        my_format = getWkbFormats(workbook, "format_lt_grey")
        my_format.set_align('center')
    else:
        # $logger->error("\nformat '$format_val' not defined");
        exit

    return (my_format)


def writeMergedHeaders(workbook, worksheet, worksheet_db, row, col, printhash):
    """Write the merged table headers to Excel file"""

    for heading in printhash['headers_merged']:
        if heading is "":
            worksheet.write(row, col, "")
            worksheet_db.write(row, col, "")
            col += 1
        else:
            num_cols = printhash['headers_merged_num'][heading]
            col_letter_first = ntol(col + 1)
            col_letter_last = ntol(col + num_cols)

            col = col + num_cols

            my_formats = getWkbFormats(workbook, "format_merged")
            merge_cells = "{}1:{}1".format(col_letter_first, col_letter_last)
            worksheet.merge_range(merge_cells, heading, my_formats)
            worksheet_db.merge_range(merge_cells, heading, my_formats)

    row += 1

    return(row, col)


def writeNonMergedHeaders(myPA, workbook, worksheet,
                          worksheet_db, row, col, printhash):
    """Write non-merged headers to Excel file"""

    # get the columns to print.  column order is from printArray
    # prepend "r_num" to array
    headingsMergedArray = ["r_num"] + \
        myPA.getHeadingsFromPrintArray(1, 1)
    headingsNotMergedArray = ["r_num"] + \
        myPA.getHeadingsFromPrintArray(0, 1)

    # print("\n", headingsNotMergedArray)

    d_headingsNotMergedArray = deque(headingsNotMergedArray)
    # print("\n", d_headingsNotMergedArray)

    for heading in headingsMergedArray:

        headingNotMerged = d_headingsNotMergedArray.popleft()

        # ($heading_new = $heading) =~ s/.*://;
        replace_val = '.*' + config['defaults_static']['MERGED_SEPARATOR']
        # print(replace)
        heading_new = heading
        heading_new = re.sub(replace_val, '', heading_new)
        # print("heading={},  headingNotMerged={}, heading_new={}".format(heading, headingNotMerged, heading_new))
        # print (row,col)
        # print("heading_new", heading_new)
        worksheet.write(row, col, heading_new,
                        getWkbFormats(workbook, "format_non_merged"))
        worksheet_db.write(row, col, heading_new,
                           getWkbFormats(workbook, "format_non_merged"))

        comment = None
        if heading is not "r_num":
            comment = \
                myPA.getSecondValueFromHeadingValueCombo(
                    "wkst_header", headingNotMerged, "comment")

        # write comments
        if comment:
            worksheet.write_comment(row, col, comment)

        col += 1

    row += 1

    return(row, col)

#  http://www.lemoda.net/perl/excel_column/excel-column.html
# Just for testing. Save the output as test.csv and open in Excel.
###############################################################################
#
# Functions used for Autofit.
#
###############################################################################

# Convert a number from 1 to 26 into a letter from A to Z.


def _number_to_letter(number):
    # if number > 26:
    #     die "Number $number out of range";
    letter = chr(number+ord("A"))
    return letter

# Convert a number into an Excel column name.


def ntol(n):
    """Convert excel column letter to a number"""
    # The maximum number of columns in an Excel worksheet is ...
    max_excel_column = 256

    if n >= 1 and n <= 26:
        l = _number_to_letter(n - 1)
    elif n > 26 and n <= max_excel_column:
        chr1 = _number_to_letter(int((n - 1)/26 - 1))
        chr2 = _number_to_letter((n - 1) % 26)
        l = chr1 + chr2
    else:
        pass
        # die "number $n out of Excel's column range (1-$max_excel_column)";
    return l


def getWkbFormats(workbook, format_request):
    """Set the cell formatting based on format_request argument"""

    if format_request is "format_default":  # default
        my_format = workbook.add_format(
            {'align': 'right'},
            )
        # my_format = workbook.add_format()
        # my_format.set_align('right')
    elif format_request is "format_grn":  # light green
        my_format = workbook.add_format(
            {'bg_color': '#C6EFCE',
             'color': '#006100'},
            )
    elif format_request is "format_lt_grey":  # light green
        my_format = workbook.add_format(
            {'bg_color': '#C0C0C0'},
            )
    elif format_request is "format_orange":  # light orange
        my_format = workbook.add_format(
            {'bg_color': '#FFCC00',
             'color': '#000000'},
            )
    elif format_request is "format_red":  # light red
        my_format = workbook.add_format(
            {'bg_color':  '#FF99CC',
             'color':     '#9C0006'},
            )
    elif format_request is "format_merged":  # merged
        my_format = workbook.add_format(
            {'bg_color': '#808080',
             'bold': 1,
             'align': 'center',
             'border': 1},
            )
    elif format_request is "format_non_merged":
        my_format = workbook.add_format(
            {'bg_color': '#C0C0C0',
             'align': 'center',
             'border': 1},
            )
    elif format_request is "format_mobilehome":
        my_format = workbook.add_format(
            {'bg_color': '#C6EFCE',
             'align': 'center',
             'border': 1},
            )

    return(my_format)
